import openpyxl
path="C:\\Users\\HP\\Desktop\\PROGRAMMING\\ab.xlsx"
wb_onj=openpyxl.load_workbook(path)
sheet_obj=wb_onj.active

for sheet_name in wb_onj.sheetnames:
    sheet=wb_onj[sheet_name]
    print(f"Title={sheet.title}")

    for value in sheet.iter_rows(values_only=True):
        print(value)

data=(("***","Thank You","***"),)

for row in data:
    sheet_obj.append(row)

wb_onj.save("C:\\Users\\HP\\Desktop\\PROGRAMMING\\ab.xlsx")
